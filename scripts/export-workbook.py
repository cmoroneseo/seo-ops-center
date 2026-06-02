#!/usr/bin/env python3
"""
export-workbook.py — Stage 3, step 1 (extract).

Reads "Sheet App Scrips/SEO Ops Command Center.xlsx" and writes a normalized
scripts/workbook-data.json that scripts/import-workbook.ts loads into Supabase.

This step ONLY parses the messy spreadsheet (cached formula values, Excel serial
dates, free-text statuses) into clean raw rows grouped by client_slug. All
business mapping (status lifecycle, slug as FK join, etc.) happens in the TS
loader so there is a single source of truth for that logic (lib/seo-ops-logic.ts).

Usage:
    python3 scripts/export-workbook.py
    python3 scripts/export-workbook.py --workbook "path/to/file.xlsx"

Requires: openpyxl  (pip install openpyxl)
"""
import argparse
import datetime as dt
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

EXCEL_EPOCH = dt.date(1899, 12, 30)  # Excel/Sheets serial-date origin


def slugify(name) -> str:
    return re.sub(r"[^a-z0-9]", "", str(name or "").lower())


def to_iso_date(v):
    """Excel serial / datetime -> 'YYYY-MM-DD', else None."""
    if v is None or v == "":
        return None
    if isinstance(v, (dt.datetime, dt.date)):
        return (v.date() if isinstance(v, dt.datetime) else v).isoformat()
    if isinstance(v, (int, float)):
        # plausible serial range (~1982..2065); ignore tiny numbers
        if 30000 <= v <= 60000:
            return (EXCEL_EPOCH + dt.timedelta(days=int(v))).isoformat()
    return None


def num(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip())
    except (ValueError, TypeError):
        return None


def s(v):
    if v is None:
        return None
    t = str(v).strip()
    return t or None


def id_str(v):
    """Render an identifier that may have been read as a float ('123.0') cleanly."""
    if v is None or v == "":
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return str(v).strip() or None


def header_map(ws, header_row=1):
    """Map normalized header text -> 0-based column index."""
    m = {}
    for i, cell in enumerate(ws[header_row]):
        h = s(cell.value)
        if h:
            m[h.strip().lower()] = i
    return m


def find(hm, *names):
    """First matching column index for any of the given (substring) names."""
    for n in names:
        n = n.lower()
        for h, idx in hm.items():
            if h == n or n in h:
                return idx
    return None


def rows_from(ws, start_row):
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if any(c is not None and str(c).strip() != "" for c in row):
            yield row


def cell(row, idx):
    return row[idx] if idx is not None and idx < len(row) else None


# --- status mapping (client level) -----------------------------------------
def map_client_status(raw):
    r = (raw or "").strip().lower()
    if r == "active":
        return "active"
    if r in ("cancelled", "canceled", "churned", "inactive"):
        return "inactive"
    return "pending"


def export(workbook_path: Path) -> dict:
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = {ws.title: ws for ws in wb.worksheets}
    out = {
        "clients": [], "deliverables": [], "time_logs": [], "monthly_plans": [],
        "client_change_log": [], "team_bonus": [], "analytics_map": [],
    }
    warnings = []

    # ---- Client Overview --------------------------------------------------
    ws = sheet.get("Client Overview")
    if ws:
        hm = header_map(ws, 1)
        c = lambda r, *n: cell(r, find(hm, *n))
        known = set(find(hm, x) for x in [
            "client", "launch date", "seo hours", "hour type", "deliverables",
            "blogs due", "account manager", "status", "tier", "target blog",
            "actual blogs", "blogs delivered", "past due", "override", "on track",
            "notes", "planning", "client id", "start month", "original launch",
            "launch date override", "first delivery",
        ] if find(hm, x) is not None)
        headers = list(hm.items())
        for r in rows_from(ws, 2):
            name = s(c(r, "client"))
            if not name:
                continue
            extra = {}
            for h, idx in headers:
                if idx not in known and idx < len(r):
                    val = s(r[idx])
                    if val is not None:
                        extra[h] = val
            out["clients"].append({
                "name": name,
                "client_slug": slugify(name),
                "launch_date": to_iso_date(c(r, "launch date")),
                "original_launch_date": to_iso_date(c(r, "original launch")),
                "launch_date_override": to_iso_date(c(r, "launch date override")),
                "seo_hours": num(c(r, "seo hours")),
                "hour_type": s(c(r, "hour type")),
                "deliverables_spec": s(c(r, "deliverables")),
                "blogs_due_per_month": num(c(r, "blogs due")),
                "account_manager_name": s(c(r, "account manager")),
                "tier": num(c(r, "tier")),
                "target_blog_count": num(c(r, "target blog")),
                "delivered_override": num(c(r, "override")),
                "notes": s(c(r, "notes")),
                "planning_tags": s(c(r, "planning")),
                "status_raw": s(c(r, "status")),
                "status": map_client_status(s(c(r, "status"))),
                "custom_fields": extra,
            })

    # ---- Client Campaigns (campaign engagement details) -------------------
    ws = sheet.get("Client Campaigns")
    campaigns = {}
    if ws:
        hm = header_map(ws, 1)
        c = lambda r, *n: cell(r, find(hm, *n))
        for r in rows_from(ws, 2):
            name = s(c(r, "client"))
            if not name:
                continue
            campaigns[slugify(name)] = {
                "total_blogs": num(c(r, "total blogs")),
                "total_hours": num(c(r, "total seo hours", "total hours")),
                "start": to_iso_date(c(r, "start date")),
                "end": to_iso_date(c(r, "end date")),
            }

    # ---- Client Analytics Map --------------------------------------------
    ws = sheet.get("Client Analytics Map")
    if ws:
        hm = header_map(ws, 1)
        c = lambda r, *n: cell(r, find(hm, *n))
        for r in rows_from(ws, 2):
            name = s(c(r, "client"))
            if not name:
                continue
            out["analytics_map"].append({
                "client_slug": slugify(name),
                "ga4_property_id": id_str(c(r, "ga4")),
                "gsc_url": s(c(r, "gsc")),
            })

    # attach campaign info onto clients
    for cl in out["clients"]:
        camp = campaigns.get(cl["client_slug"])
        ht = (cl.get("hour_type") or "").strip().lower()
        spec = (cl.get("deliverables_spec") or "").lower()
        is_campaign = ht == "campaign" or "campaign:" in spec or camp is not None
        cl["engagement_model"] = "Campaign" if is_campaign else "Retainer"
        cl["campaign"] = camp

    # ---- Deliverables Tracker --------------------------------------------
    ws = sheet.get("Deliverables Tracker")
    if ws:
        hm = header_map(ws, 1)
        c = lambda r, *n: cell(r, find(hm, *n))
        for r in rows_from(ws, 2):
            name = s(c(r, "client"))
            title = s(c(r, "blog title", "title"))
            if not name or not title:
                continue
            due = to_iso_date(c(r, "due date"))
            month_serial = to_iso_date(c(r, "month"))
            month = (month_serial or due or "")[:7] or None
            out["deliverables"].append({
                "client_slug": slugify(name),
                "title": title,
                "type_raw": s(c(r, "deliverables")),  # e.g. "Blog"
                "status_raw": s(c(r, "status")),
                "delivered_on_raw": s(c(r, "delivered on")),
                "due_date": due,
                "month": month,
                "notes": s(c(r, "notes")),
                "account_manager_name": s(c(r, "account manager")),
            })

    # ---- Daily Hours Log --------------------------------------------------
    ws = sheet.get("Daily Hours Log")
    if ws:
        hm = header_map(ws, 1)
        c = lambda r, *n: cell(r, find(hm, *n))
        for r in rows_from(ws, 2):
            date = to_iso_date(c(r, "date"))
            name = s(c(r, "client"))
            hours = num(c(r, "hours"))
            if not date or not name or hours is None:
                continue
            out["time_logs"].append({
                "date": date,
                "client_slug": slugify(name),
                "account_manager_name": s(c(r, "account manager")),
                "hours": hours,
                "description": s(c(r, "task", "notes")),
            })

    # ---- Monthly Planners (every "<Month> YYYY Planner" sheet) -----------
    planner_re = re.compile(r"^([A-Za-z]+)\s+(\d{4})\s+Planner$")
    for title, ws in sheet.items():
        m = planner_re.match(title)
        if not m:
            continue
        try:
            month = dt.datetime.strptime(f"{m.group(1)} {m.group(2)}", "%B %Y").strftime("%Y-%m")
        except ValueError:
            warnings.append(f"Could not parse planner month from '{title}'")
            continue
        hm = header_map(ws, 2)  # planners: row 1 = week dates, row 2 = headers
        c = lambda r, *n: cell(r, find(hm, *n))
        for r in rows_from(ws, 3):
            name = s(c(r, "client name", "client"))
            if not name:
                continue
            weeks = []
            for w in range(1, 6):
                planned = num(c(r, f"w{w} planned"))
                logged = num(c(r, f"w{w} logged"))
                variance = num(c(r, f"w{w} variance"))
                if planned is None and logged is None and variance is None:
                    continue
                weeks.append({"week": w, "planned": planned, "logged": logged, "variance": variance})
            if not weeks:
                continue
            out["monthly_plans"].append({
                "client_slug": slugify(name),
                "month": month,
                "weeks": weeks,
                "notes": s(c(r, "notes")),
            })

    # ---- Client Change Log -----------------------------------------------
    ws = sheet.get("Client Change Log")
    if ws:
        hm = header_map(ws, 1)
        c = lambda r, *n: cell(r, find(hm, *n))
        for r in rows_from(ws, 2):
            name = s(c(r, "client"))
            if not name:
                continue
            out["client_change_log"].append({
                "client_slug": slugify(name),
                "date_of_change": to_iso_date(c(r, "date of change")),
                "changed_by": s(c(r, "changed by")),
                "prev_seo_hours": num(c(r, "previous seo hours", "prev seo")),
                "new_seo_hours": num(c(r, "new seo hours")),
                "prev_blog_count": num(c(r, "previous blog", "prev blog")),
                "new_blog_count": num(c(r, "new blog")),
                "effective_date": to_iso_date(c(r, "effective date")),
                "notes": s(c(r, "notes")),
            })

    # ---- Team Bonus Tracker (best-effort) --------------------------------
    ws = sheet.get("SEO Team Bonus Tracker")
    if ws:
        members = {slugify(cl["account_manager_name"]): cl["account_manager_name"]
                   for cl in out["clients"] if cl.get("account_manager_name")}
        month = None
        title = s(ws.cell(1, 1).value) or ""
        mm = re.search(r"([A-Za-z]+)\s+(\d{4})", title)
        if mm:
            try:
                month = dt.datetime.strptime(f"{mm.group(1)} {mm.group(2)}", "%B %Y").strftime("%Y-%m")
            except ValueError:
                pass
        for r in rows_from(ws, 1):
            name = s(r[0]) if r else None
            if not name or slugify(name) not in members:
                continue
            total = next((num(x) for x in r[1:6] if num(x) is not None), None)
            if total is None:
                continue
            out["team_bonus"].append({
                "member_name": name, "month": month,
                "base_from_hours": total, "kpi_bonus": 0, "cap": 300,
                "notes": "Imported (best-effort) from SEO Team Bonus Tracker",
            })
        if not out["team_bonus"]:
            warnings.append("Team bonus rows not auto-detected — review the sheet layout / enter manually.")

    out["_meta"] = {
        "source": str(workbook_path),
        "counts": {k: len(v) for k, v in out.items() if isinstance(v, list)},
        "warnings": warnings,
    }
    return out


def main():
    ap = argparse.ArgumentParser()
    repo_root = Path(__file__).resolve().parent.parent
    ap.add_argument("--workbook", default=str(repo_root / "Sheet App Scrips" / "SEO Ops Command Center.xlsx"))
    ap.add_argument("--out", default=str(repo_root / "scripts" / "workbook-data.json"))
    args = ap.parse_args()

    wb_path = Path(args.workbook)
    if not wb_path.exists():
        sys.exit(f"Workbook not found: {wb_path}")

    data = export(wb_path)
    Path(args.out).write_text(json.dumps(data, indent=2, ensure_ascii=False))
    print(f"Wrote {args.out}")
    print("Counts:", json.dumps(data["_meta"]["counts"], indent=2))
    if data["_meta"]["warnings"]:
        print("Warnings:")
        for w in data["_meta"]["warnings"]:
            print("  -", w)


if __name__ == "__main__":
    main()
